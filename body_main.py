'''
Часть 1:
Приведение столбца из книги book_2 в идентичный формат с book_1,
для их последующего сравнения
Даты в book_1 - дата из выгрузки системы, файл выгрузки с расширением .xlsx
Даты в book_2 - фактическая дата приема препарата, заполненная вручную, имеет ошибки
модули re, OpenPyXL
'''


import openpyxl
from openpyxl.styles import PatternFill
import re

wb = openpyxl.load_workbook("book_2.xlsx") #workbook book_2, данная книга с некорректным форматом даты
ws = wb["sample_take"] #worklist 'sample_take'

class Replacement:
    def tabulation():
        counter = 0
        for r in range(1, ws.max_row + 1): #r row
            for c in range(1, ws.max_column + 1): #c column
                s = ws.cell(r, c).value
                if s is not None and isinstance(s, str): #проверка на пустое значение
                    ws.cell(r, c).value = re.sub(r'[^a-zA-Z0-9]', ",", ws.cell(r, c).value)
                print("row {} col {} : {}".format(r, c, s))
                counter += 1

        wb.save('compared.xlsx')
        print("изменено ячеек {}, исправлена дата".format(counter))

Replacement.tabulation()

'''
Часть 2:
Сравнение двух файлов и выделение цветом несоответствий
модуль OpenPyXL 
'''
data_compared = openpyxl.load_workbook("compared.xlsx")
data_main = openpyxl.load_workbook("book_3.xlsx")

fill_st = PatternFill(start_color="FFBF00", end_color="FBCEB1", fill_type="solid")

ws_res_orig = data_main['sample_orig']
ws_res_comp = data_compared['sample_take']

def comparison():
    for row in ws_res_comp.iter_rows():
        for cell in row:
            cell_value = cell.value
            cell_loc = cell.coordinate

        if cell_value != ws_res_orig[cell_loc].value:
            cell.fill = fill_st

comparison()
data_compared.save("result.xlsx")